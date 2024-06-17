# BD 2023-24
# This class is intended to serve as a GUI display manager for the body forces simulation research project.

## TODO:
#   - fix sync issues with root.mainloop blocking proper updating of the manual calibration toggle
#       when calling function from outside the main program loop
#       - basically, by changing the value of self.manual_calibration from a function call that
#           didn't originate from the main loop (i.e. from within root.mainloop()), the update
#           attempt is blocked by root.mainloop()


import numpy as np
import math
from matplotlib import pyplot as plt
import time
import threading
import os

import cv2

from tkinter import *
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image
from PIL import ImageTk

import livestream_mediapipe_class as lsmp   # custom class, handles mediapipe

# for use recording to excel doc
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


### OPTIONS

# model to use for mediapipe
pose_landmarker = 'landmarkers/pose/pose_landmarker_full.task'
hand_landmarker = 'landmarkers/hand/hand_landmarker.task'
face_landmarker = 'landmarkers/face/face_landmarker.task'

# check if windows; if so, use windows file pathing
if os.name == 'nt':
    pose_landmarker = "landmarkers\\pose\\pose_landmarker_full.task"
    hand_landmarker = "landmarkers\\hand\\hand_landmarker.task"
    face_landmarker = 'landmarkers\\face\\face_landmarker.task'


# load and prep placeholder image for program initialization
no_image_path = './no_image.png'            # placeholder image location
#no_image = Image.fromarray(cv2.cvtColor(cv2.imread(no_image_path), cv2.COLOR_BGR2RGB))


# setup runnable class for management of the GUI
class SimGUI():

    # initialization
    def __init__(self) -> None:
        
        ### DATA AND CONSTANTS

        # variable for dynamic width of settings
        self.settings_width = 20

        # set up dictionary to read from for gui display of data (bicep force data and elbow angles)
        self.calculated_data = {
            "right_bicep_force": "NaN",
            "right_elbow_angle": "NaN",
            "left_bicep_force": "NaN",
            "left_elbow_angle": "NaN",
            "uarm_spher_coords": "NaN",#["NaN", "NaN", "NaN"],
            "farm_spher_coords": "NaN"#["NaN", "NaN", "NaN"]
        }

        # dict for containing hand data
        self.hand_data = np.zeros((2, 2), dtype = "float32")     # [left[epsilon, phi], right[epsilon, phi]]

        # store past bicep force calculations
        self.history_bicep_force = np.ndarray((1), dtype = "float32")
        self.history_elbow_angle = np.ndarray((1), dtype = "float32")
        self.hbf_max_len = 1000             # max length for history of bicep force

        # initialize mediapipe thread
        self.mediapipe_runtime = lsmp.Pose_detection(pose_landmarker, hand_landmarker, face_landmarker)
        self.mediapipe_runtime.start()

        # allow entry in imperial (instead of metric)
        self.use_imperial = False

        # allow auto update of graph (WARNING: lags current setup)
        self.auto_update_graph = False

        # toggle manual conversion/calibration ratio/factor
        self.manual_calibration = False


        ### EXCEL DATA RECORDING

        # name of excel file
        self.xl_filename = "angles_p-up.xlsx"
        # data to record
        #   must be an array
        #   make sure to update in `update_data()`
        self.xl_desired_data = [self.hand_data[0, 0]]     # phi for left hand

        # start row of spreadsheet
        self.xl_start_row = 5   # starting at 5 to give room for things like avg, std dev, and std err
        # current column of spreadsheet
        self.xl_cur_col = 2
        # current row of spreadsheet
        self.xl_cur_row = 5
        # whether or not data is being recorded
        self.xl_is_recording = False
        # time at which current recording should end
        self.xl_cur_end_time = 0
        # length (in seconds) of recording trials
        self.xl_trial_length = 10

        # initialize excel spreadsheet
        self.workbook = Workbook()
        self.xl_spreadsheet = self.workbook.active

        # set description for first few rows
        self.xl_spreadsheet.cell(row = 1, column = 1).value = "Avg: "
        self.xl_spreadsheet.cell(row = 2, column = 1).value = "Std dev: "
        self.xl_spreadsheet.cell(row = 3, column = 1).value = "Std err: "


        ### GUI SETUP

        # delay between frames
        self.delay = 15

        # data update interval
        self.update_interval = 200      # update every 200 milliseconds

        # image height/width
        self.height, self.width = self.mediapipe_runtime.get_height_width()
        
        # initialize root of the tkinter gui display
        self.root = Tk()
        self.root.title("Biomechanics Simulation")

        # create frame
        self.gui = Frame(self.root)#, bg = "white")
        self.gui.grid()

        # create image label in frame
        self.image_label = Label(self.gui)
        self.image_label.grid(row = 0, column = 0)
        self.image_label.photo = None


        ### GUI ORGANIZATION 

        # grid section for containing all textual info
        self.data = Frame(self.gui)
        self.data.grid(row = 0, column = 1)

        # grid section for settings/calibration
        self.settings = LabelFrame(self.data, text = "Settings:")
        self.settings.grid(row = 0, column = 0)

        # settings for unit conversion factor (metric <--> sim units)
        self.ucf_label = Label(self.settings, text = "Unit conversion factor: ", height = 1, width = self.settings_width)
                               #cursor = "Approximate conversion ratio between metric units and simulation units.\n" + 
                               #         "Only intended for use with \"Manual\" functionality.")
        self.ucf_var = StringVar()
        self.ucf_entry = Entry(self.settings, textvariable = self.ucf_var)
        self.ucf_toggle_var = IntVar()
        self.ucf_toggle = Checkbutton(self.settings, text = "Manual", variable = self.ucf_toggle_var,       # now defunct, does effectively nothing with new calibration system
                                      onvalue = 1, offvalue = 0, height = 1, width = 10, command = self.toggle_manual_conversion)
        self.ucf_submit = Button(self.settings, text = "Submit", command = self.set_conversion_ratio)
        self.ucf_label.grid(row = 1, column = 0)
        self.ucf_entry.grid(row = 1, column = 1)
        self.ucf_toggle.grid(row = 2, column = 0)
        self.ucf_submit.grid(row = 2, column = 1)

        # biacromic scale factor
        self.bsf_scale = Scale(self.settings, from_ = 0.22, to = 0.24, orient = "horizontal", length = 200, 
                                label = "Biacromic (shoulder width) Scale", showvalue = True, command = self.set_bsf, resolution = 0.001)
        self.bsf_scale.grid(row = 3, columnspan = 2)

        # allow entry in imperial unit system (as opposed to metric)
        self.ms_label = Label(self.settings, text = "Use imperial system? (Default: metric)")
        self.ms_var = IntVar()
        self.ms_toggle = Checkbutton(self.settings, variable = self.ms_var, onvalue = 1, offvalue = 0, command = self.toggle_imperial)
        self.ms_label.grid(row = 4, column = 0)
        self.ms_toggle.grid(row = 4, column = 1)

        # allow adjustment of image height and width
        #self.image_adjust = LabelFrame(self.settings, text = "Image adjustment:")   # set up lavel frame to section off this part of the settings
        #self.image_adjust.grid(row = 5, column = 0)
        # height
        #self.image_height_label = Label(self.image_adjust, text = "Image height: ", height = 1, width = self.settings_width)
        #self.image_height_var = StringVar()
        #self.image_height_var.set("640")
        #self.image_height_entry = Entry(self.image_adjust, textvariable = self.image_height_var)
        #self.image_height_label.grid(row = 1, column = 0)
        #self.image_height_entry.grid(row = 1, column = 1) 
        # width
        #self.image_width_label = Label(self.image_adjust, text = "Image width: ", height = 1, width = self.settings_width)
        #self.image_width_var = StringVar()
        #self.image_width_var.set("480")
        #self.image_width_entry = Entry(self.image_adjust, textvariable = self.image_width_var)
        #self.image_width_label.grid(row = 2, column = 0)
        #self.image_width_entry.grid(row = 2, column = 1) 
        #submit button
        #self.submit_image_hw = Button(self.image_adjust, text = "Submit", command = self.set_livestream_hw)
        #self.submit_image_hw.grid(row = 3, column = 1)



        # grid section for user input
        self.user_input = LabelFrame(self.data, text = "User input:")
        self.user_input.grid(row = 1, column = 0)

        # height user input
        self.height_label = Label(self.user_input, text = "User height (meters): ", height = 1, width = self.settings_width)
        self.height_var = StringVar()
        self.height_var.set("1.78")                    # set to default value
        self.height_entry = Entry(self.user_input, textvariable = self.height_var)
        self.height_label.grid(row = 1, column = 0)
        self.height_entry.grid(row = 1, column = 1)

        # weight user input
        self.weight_label = Label(self.user_input, text = "User weight (kilograms): ", height = 1, width = self.settings_width)
        self.weight_var = StringVar()
        self.weight_var.set("90")                    # set to default value
        self.weight_entry = Entry(self.user_input, textvariable = self.weight_var)
        self.weight_label.grid(row = 2, column = 0)
        self.weight_entry.grid(row = 2, column = 1)

        # ball mass user input
        self.bm_label = Label(self.user_input, text = "Ball mass (kilograms): ", height = 1, width = self.settings_width)
        self.bm_var = StringVar()
        self.bm_var.set("3")                    # set to default value
        self.bm_entry = Entry(self.user_input, textvariable = self.bm_var)
        self.bm_label.grid(row = 3, column = 0)
        self.bm_entry.grid(row = 3, column = 1)

        # button to submit height and weight and ball mass
        self.submit_hw = Button(self.user_input, text = "Submit", command = self.hw_submit)
        self.submit_hw.grid(row = 4, column = 1)

        
        ## excel data output: button to start recording data to excel document, as well as status of recording
        # grid section for excel
        self.xl_input = LabelFrame(self.data, text = "Record desired data to excel: ")
        self.xl_input.grid(row = 2, column = 0)
        # status of recording
        self.xl_status_var = StringVar()
        self.xl_status_var.set("Press \"Start\" to begin")
        # current trial number
        self.xl_cur_trial_var = StringVar()
        self.xl_cur_trial_var.set("1")
        # button to start recording
        self.xl_start_rec_button = Button(self.xl_input, text = "Start", command = self.xl_start_rec)
        # current trial number
        self.xl_cur_trial = Label(self.xl_input, textvariable = self.xl_cur_trial_var, height = 1, width = int(self.settings_width / 2))
        # label for current trial
        self.xl_cur_trial_label = Label(self.xl_input, text = "Current trial: ", height = 1, width = int(self.settings_width / 2))
        # current status of recording
        self.xl_status = Label(self.xl_input, textvariable = self.xl_status_var, height = 1, width = self.settings_width)
        # grid the excel gui elements
        self.xl_status.grid(row = 0)
        self.xl_start_rec_button.grid(row = 1, column = 1)
        self.xl_cur_trial_label.grid(row = 2, column = 0)
        self.xl_cur_trial.grid(row = 2, column = 1)



        # grid section for data output
        self.data_output = LabelFrame(self.data, text = "Data output:")
        self.data_output.grid(row = 3, column = 0)

        # RIGHT ARM:
        self.do_right = LabelFrame(self.data_output, text = "Right arm:")
        self.do_right.grid(row = 1, column = 0)
        # bicep force output
        self.right_bicep_label = Label(self.do_right, text = "Bicep force (Newtons): ", height = 1, width = self.settings_width)
        self.right_bicep_var = StringVar()
        self.right_bicep_var.set(str(self.calculated_data["right_bicep_force"]))
        self.right_bicep_force = Label(self.do_right, textvariable = self.right_bicep_var, height = 1, width = int(self.settings_width / 2), relief = GROOVE)
        self.right_bicep_label.grid(row = 1, column = 0)
        self.right_bicep_force.grid(row = 1, column = 1)

        # elbow angle output
        self.right_elbow_label = Label(self.do_right, text = "Elbow angle (Degrees): ", height = 1, width = self.settings_width)
        self.right_elbow_var = StringVar()
        self.right_elbow_var.set(str(self.calculated_data["right_elbow_angle"]))
        self.right_elbow_angle = Label(self.do_right, textvariable = self.right_elbow_var, height = 1, width = int(self.settings_width / 2), relief = GROOVE)
        self.right_elbow_label.grid(row = 2, column = 0)
        self.right_elbow_angle.grid(row = 2, column = 1)

        # LEFT ARM:
        self.do_left = LabelFrame(self.data_output, text = "Left arm:")
        self.do_left.grid(row = 2, column = 0)
        # bicep force output
        self.left_bicep_label = Label(self.do_left, text = "Bicep force (Newtons): ", height = 1, width = self.settings_width)
        self.left_bicep_var = StringVar()
        self.left_bicep_var.set(str(self.calculated_data["left_bicep_force"]))
        self.left_bicep_force = Label(self.do_left, textvariable = self.left_bicep_var, height = 1, width = int(self.settings_width / 2), relief = GROOVE)
        self.left_bicep_label.grid(row = 1, column = 0)
        self.left_bicep_force.grid(row = 1, column = 1)

        # elbow angle output
        self.left_elbow_label = Label(self.do_left, text = "Elbow angle (Degrees): ", height = 1, width = self.settings_width)
        self.left_elbow_var = StringVar()
        self.left_elbow_var.set(str(self.calculated_data["left_elbow_angle"]))
        self.left_elbow_angle = Label(self.do_left, textvariable = self.left_elbow_var, height = 1, width = int(self.settings_width / 2), relief = GROOVE)
        self.left_elbow_label.grid(row = 2, column = 0)
        self.left_elbow_angle.grid(row = 2, column = 1)



        # set up bicep force scatter plot
        self.fig, self.ax = plt.subplots()
        #self.hist_bf_plot = self.ax.scatter(self.history_elbow_angle, self.history_bicep_force)
        self.ax.set_ylim(ymin = 0, ymax = 1000)
        self.ax.set_xlim(xmin = 0, xmax = 180)
        self.ax.set_xlabel("Elbow angle")
        self.ax.set_ylabel("Bicep force")
        self.ax.set_title("Bicep force vs Elbow angle (left arm)")

        # setup plot in tkinter gui (alongside button to update it)
        self.forces_graph = FigureCanvasTkAgg(self.fig, master = self.gui)
        self.forces_graph_widget = self.forces_graph.get_tk_widget()
        self.forces_graph_widget.grid(row = 1, column = 0)
        self.forces_graph_grid = LabelFrame(self.gui, text = "Graph settings:")
        self.forces_graph_grid.grid(row = 1, column = 1)
        self.forces_graph_update = Button(self.forces_graph_grid, text = "Update graph", command = self.update_scatterplot)
        self.forces_graph_au_var = IntVar()
        self.forces_graph_autoupdate = Checkbutton(self.forces_graph_grid, text = "Auto update graph", variable = self.forces_graph_au_var, onvalue = 1, offvalue = 0, command = self.toggle_graph_autoupdate)
        self.forces_graph_update.grid(row = 0, column = 0)
        self.forces_graph_autoupdate.grid(row = 1, column = 0)



    
    # start/run the gui display
    def start(self):
        # start updater loops
        self.update_display()                               # update display
        self.update_data()                                  # update numerical data
        #self.mediapipe_runtime.run()
        #self.root.update_idletasks()
        #self.gui.update_idletasks()

        # handle program close
        self.root.protocol("WM_DELETE_WINDOW", self.__del__)

        # allow closing program by pressing escape
        self.root.bind("<Escape>", lambda event: self.__del__(event))

        # start the display
        self.root.mainloop()


    # update the data being displayed
    def update_display(self):#, new_frame, data_dict):
        # handle frame/image data
        ret, frame = self.mediapipe_runtime.get_cur_frame()
        frame = cv2.cvtColor(cv2.flip(frame,1), cv2.COLOR_BGR2RGB)      # converting back to RGB for display

        if ret:                                             # only update if frame is present
            self.image_label.photo = ImageTk.PhotoImage(image = Image.fromarray(frame))
            self.image_label.configure(image = self.image_label.photo)
            self.calculated_data = self.mediapipe_runtime.ep.get_calculated_data()
            self.hand_data = self.mediapipe_runtime.ep.get_hand_data()

        # call next update cycle
        self.root.after(self.delay, self.update_display)    # update approximately 60 times per second

    # update numerical data in gui
    def update_data(self):
        
        # elbow angles
        self.right_elbow_var.set(str(self.calculated_data["right_elbow_angle"]))
        self.left_elbow_var.set(str(self.calculated_data["left_elbow_angle"]))
        
        # bicep forces
        # only show numerical bicep force calculations if elbow angle > 90 degrees (otherwise, formula doesn't work)
        if (float(self.calculated_data["right_elbow_angle"]) >= 90):    # right arm
            self.right_bicep_var.set(str(self.calculated_data["right_bicep_force"]))
        else:
            self.right_bicep_var.set("N/A (angle below 90 deg)")
        if (float(self.calculated_data["left_elbow_angle"]) >= 90):    # left arm
            self.left_bicep_var.set(str(self.calculated_data["left_bicep_force"]))
        else:
            self.left_bicep_var.set("N/A (angle below 90 deg)")

        # update manual calibration
        self.manual_calibration = self.mediapipe_runtime.toggle_auto_calibrate
        # check if using manual calibration
        if not self.manual_calibration:
            self.ucf_var.set(str("%0.5f" % self.mediapipe_runtime.ep.get_conversion_ratio()))

        # update elbow angle and bicep force data
        self.update_bicep_array()
        # optional live plot updater
        if self.auto_update_graph:
            self.update_scatterplot()

        # handle excel recording after data is updated
        self.xl_desired_data = [self.hand_data[0, 0]]  # left hand phi     # update desired_data
        self.xl_update()

        # call next update cycle
        self.gui.after(self.update_interval, self.update_data)


    # handle keeping track of the past n timesteps of (left arm) body force calculations
    def update_bicep_array(self):
        # if above certain n value, remove the oldest data before adding the newest
        if (np.shape(self.history_bicep_force)[0] >= self.hbf_max_len):
            np.delete(self.history_bicep_force, 0)
            np.delete(self.history_elbow_angle, 0)  # assume same is true for elbow data
        
        # append newest data to array for use by matplotlib to display graph
        self.history_bicep_force = np.append(self.history_bicep_force, float(self.calculated_data["left_bicep_force"]))
        self.history_elbow_angle = np.append(self.history_elbow_angle, float(self.calculated_data["left_elbow_angle"]))

    # update scatterplot of bicep forces vs elbow angle
    def update_scatterplot(self):
        # update plot
        self.ax.scatter(self.history_elbow_angle, self.history_bicep_force)
        self.fig.canvas.draw()


    ### USER INPUT HANDLERS

    # height and weight submission
    def hw_submit(self):
        height = float(self.height_entry.get())
        weight = float(self.weight_entry.get())
        ball = float(self.bm_entry.get())

        # convert from imperial to metric (if used)
        if self.use_imperial:
            height = height / 3.28084   # feet to meters
            weight = weight / 2.2046    # pounds to kilograms
            ball = ball / 2.2046        # pounds to kilograms

        self.mediapipe_runtime.ep.set_hwb(height, weight, ball)

    # handle toggleable measurement system
    def toggle_imperial(self):
        toggle = bool(self.ms_var.get())

        self.use_imperial = toggle

        # change GUI to accomodate
        if toggle:
            self.height_label.config(text = "User height (feet): ")
            self.weight_label.config(text = "User weight (pounds): ")
            self.bm_label.config(text = "Ball mass (pounds): ")
        # if untoggled, return to default
        else:
            self.height_label.config(text = "User height (meters): ")
            self.weight_label.config(text = "User weight (kilograms): ")
            self.bm_label.config(text = "Ball mass (kilograms): ")
    
    # handle toggle of auto graph update
    def toggle_graph_autoupdate(self):
        toggle = bool(self.forces_graph_au_var)

        self.auto_update_graph = toggle

    # handle togglable manual conversion factor/ratio
    def toggle_manual_conversion(self):
        toggle = bool(self.ucf_toggle_var.get())

        self.mediapipe_runtime.toggle_auto_conversion(toggle)

    # set manual conversion factor/ratio
    def set_conversion_ratio(self):
        ratio = float(self.ucf_entry.get())

        self.mediapipe_runtime.ep.set_conversion_ratio(ratio)

    # set biacromial (i.e. shoulder width) scale factor
    def set_bsf(self, bsf_data):
        bsf = float(bsf_data)

        self.mediapipe_runtime.ep.set_biacromial(bsf)

    # set height and width of display/camera picture view
    def set_livestream_hw(self):
        # set local variables
        self.height = int(self.image_height_var.get())
        self.width = int(self.image_width_var.get())

        # set opencv video stream/source height and width
        self.mediapipe_runtime.set_image_hw(self.height, self.width)


    ### EXCEL DATA RECORDING FUNCTIONS

    # start recording data
    def xl_start_rec(self):
        try:
            if not self.xl_is_recording:
                # set status to recording
                self.xl_is_recording = True
                # set time to end recording
                self.xl_cur_end_time = datetime.now().timestamp() + self.xl_trial_length
                # set current row to base row before trial starts
                self.xl_cur_row = self.xl_start_row
                # label current set of data being recorded
                self.xl_spreadsheet.cell(row = self.xl_cur_row, column = self.xl_cur_col).value = self.xl_cur_trial_var.get()
                # update status in gui
                self.xl_status_var.set("Recording data...")
            else:
                self.xl_status_var.set("Please wait for trial to end...")
        except Exception as e:
            print("gui.py: Exception thrown in `xl_start_rec()`\n\t%s" % str(e))

    # record current frame of desired data
    def xl_record_to_sheet(self):
        try:
            # used to reset xl_cur_col after recording data
            init_col = self.xl_cur_col

            # go to next row
            self.xl_cur_row += 1

            # iterate thru each of the desired data
            for data in self.xl_desired_data:
                # record current data
                self.xl_spreadsheet.cell(row = self.xl_cur_row, column = self.xl_cur_col).value = data
                # go to next column for recording next data
                self.xl_cur_col += 1

            # reset column to initial column
            self.xl_cur_col = init_col
        except Exception as e:
            print("gui.py: Exception thrown in `xl_record_to_sheet()`\n\t%s" % str(e))

    # update function to be called each frame when xl_is_recording is True
    def xl_update(self):
        try:
            # check if is recording
            if self.xl_is_recording:
                # check if time is up
                if (datetime.now().timestamp() < self.xl_cur_end_time):
                    self.xl_record_to_sheet()
                # end recording otherwise
                else:
                    self.xl_is_recording = False
                    # run calculations for the current run
                    self.xl_calc_err()
                    # update current trial number
                    self.xl_cur_trial_var.set(str( int(self.xl_cur_trial_var.get()) + 1 ))
                    # go to next free column
                    self.xl_cur_col += len(self.xl_desired_data)
                    # update gui status
                    self.xl_status_var.set("Done!")
                
            # update gui status after a few seconds upon completion
            elif (datetime.now().timestamp() > (self.xl_cur_end_time + 5)):
                # set status back to original status
                self.xl_status_var.set("Press \"Start\" to begin")
        except Exception as e:
            print("gui.py: Exception thrown in `xl_update()`\n\t%s" % str(e))
        
    # handle calculations at end of current excel recording run
    def xl_calc_err(self):
        try:
            # temp store for current column
            init_col = self.xl_cur_col

            # go thru for each of the desired data
            for i in range(0, len(self.xl_desired_data)):
                # current column index
                cur_col = init_col + i
                # length of current column
                cur_col_len = len(self.xl_spreadsheet[get_column_letter(cur_col)])
                # number of elements in column
                cur_len = cur_col_len - self.xl_start_row

                # get sum of current column
                cur_sum = sum(self.xl_spreadsheet.cell(row = r, column = cur_col).value for r in range(self.xl_start_row + 1, cur_col_len)) 
                # get average
                cur_avg = cur_sum / cur_len
                # get sum of squares
                #   must be done separately from cur_sum, since it uses cur_avg, which uses cur_sum
                cur_sos = sum( ( (self.xl_spreadsheet.cell(row = r, column = cur_col).value - cur_avg) ** 2 ) for r in range(self.xl_start_row + 1, cur_col_len))
                # get variance
                cur_var = cur_sos / cur_len
                # get standard deviation
                cur_stddev = np.sqrt(cur_var)
                # get standard error
                cur_stderr = cur_stddev / np.sqrt(cur_len)

                # DEBUG
                print("\nSum: %s\nAvg: %s\nStd dev: %s\nStd err: %s\n" % (cur_sum, cur_avg, cur_stddev, cur_stderr))
                
                # record in spreadsheet
                self.xl_spreadsheet.cell(row = 1, column = cur_col).value = cur_avg
                self.xl_spreadsheet.cell(row = 2, column = cur_col).value = cur_stddev
                self.xl_spreadsheet.cell(row = 3, column = cur_col).value = cur_stderr
        except Exception as e:
            print("gui.py: Exception thrown in `xl_calc_err()`\n\t%s" % str(e))



    # handle end of runtime
    def __del__(self, e = ""):
        # save excel document if any data was recorded
        if int(self.xl_cur_trial_var.get()) > 1:
            self.workbook.save(filename = self.xl_filename)

        # stop gui
        self.root.destroy()

        # stop mediapipe
        if self.mediapipe_runtime.webcam_stream.isOpened():
            self.mediapipe_runtime.webcam_stream.release()
        self.mediapipe_runtime.set_stop(True)
        self.mediapipe_runtime.join()

        # end gui
        #self.root.destroy()

if __name__ == '__main__':
    gui = SimGUI()
    gui.start()